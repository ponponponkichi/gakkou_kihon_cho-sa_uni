import os
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import time
import zipfile
import io
import datetime
import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import streamlit as st
import tempfile
import re

# --- Streamlitの基本設定 ---
st.set_page_config(page_title="学校基本調査 一括DL＆整理ツール", layout="wide")

# 列名の表記ゆれを統一するための辞書
rename_dict = {}

# 過去のログから判明している様式のマスターリスト
known_forms = {
    "07go_1": "07go_1_教員数（本務者）（再掲）", 
    "07go_A": "07go_A_学生数", 
    "07go_B": "07go_B_教員数（本務者）", 
    "07go_C": "07go_C_職員数", 
    "07go_Z": "07go_Z_教員数（兼務者）",
    "08go_2": "08go_2_学部_学科別学生数のうち休学者数", 
    "08go_3": "08go_3_学部_学科別学生数のうち最低在学年限超過学生数（編入学者は除く。）", 
    "08go_7": "08go_7_学部_専攻科，別科及び科目等履修生等の学生数", 
    "08go_D": "08go_D_学部_学科別学生数|入学志願者数|入学者数", 
    "08go_G": "08go_G_学部_出身高校の所在地県別入学者数", 
    "08go_O": "08go_O_学部_年齢別入学者数（再掲）", 
    "08go_R": "08go_R_学部_短期大学・高等専門学校・専修学校（専門課程）・高等学校等専攻科からの編入学者数",
    "09go_4": "09go_4_大学院_専攻別学生数のうち休学者数", 
    "09go_5": "09go_5_大学院_専攻別学生数のうち最低在学年限超過学生数（編入学者は除く。）", 
    "09go_8": "09go_8_大学院_科目等履修生等の学生数", 
    "09go_H": "09go_H_大学院_専攻別学生数|左記のうち社会人", 
    "09go_I": "09go_I_大学院_入学状況|入学志願者数|入学者数", 
    "09go_S": "09go_S_大学院_年齢別入学者数",
    "10go_6": "10go_6_本科学生内訳_学科別学生数のうち休学者数", 
    "10go_9": "10go_9_本科学生内訳_専攻科，別科及び科目等履修生等の学生数", 
    "10go_J": "10go_J_本科学生内訳_学科別学生数|入学状況（本科）", 
    "10go_K": "10go_K_本科学生内訳_出身高校の所在地県別入学者数", 
    "10go_Q": "10go_Q_本科学生内訳_齢別入学者数（再掲）", 
    "10go_T": "10go_T_本科学生内訳_高等学校等専攻科からの編入学者数",
    "11go_bekkei": "11go_bekkei_国費留学生，私費留学生，留学生以外の外国人学生（専攻科・別科の学生，科目等履修生・聴講生・研究生）", 
    "11go_gkssu": "11go_gkssu_国費留学生，私費留学生，留学生以外の外国人学生",
    "20go": "20go_学校施設",
    "30go_2_1": "30go_2_1_卒業後_状況別卒業者数，入学年度別卒業者数", 
    "30go_2_1_bekkei": "30go_2_1_bekkei_卒業後_年齢別卒業者数", 
    "30go_2_2": "30go_2_2_卒業後_職業別就職者数，産業別就職者数"
}

def get_hidden_header_index(filepath):
    """ファイル形式に合わせて非表示行を特定する関数"""
    if filepath.endswith('.xlsx'):
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active
            for row_idx in range(1, sheet.max_row + 1):
                if sheet.row_dimensions[row_idx].hidden:
                    return row_idx - 1
        except Exception:
            pass
    elif filepath.endswith('.xls'):
        try:
            import xlrd
            book = xlrd.open_workbook(filepath, formatting_info=True)
            sheet = book.sheet_by_index(0)
            for row_idx in range(sheet.nrows):
                if row_idx in sheet.rowinfo_map and sheet.rowinfo_map[row_idx].hidden:
                    return row_idx
        except Exception:
            pass
    return None

# --- UIの構築 ---
st.title("📊 学校基本調査 一括ダウンロード＆整理ツール")
st.write("大学改革支援・学位授与機構の「大学基本情報」から、「学校基本調査」指定様式のデータを一括取得し・複数年データを1本化します。")

# --- 追加機能：年度指定 ---
st.write("#### 📅 対象年度の指定")
years_list = ["すべて"] + [str(y) for y in range(2012, 2036)]
col1, col2 = st.columns(2)
with col1:
    start_year_str = st.selectbox("開始年度", years_list, index=0)
with col2:
    end_year_str = st.selectbox("終了年度", years_list, index=0)

st.write("#### 📂 処理モードの選択")
# ラジオボタン
mode_str = st.radio(
    "整理する範囲を選んでください",
    ("① 全様式を1本化する（5分ほどかかる）", "② 指定様式のみ1本化する（おすすめ）"),
    horizontal=True
)
current_mode = "all" if "①" in mode_str else "select"

# チェックボックス（指定様式のみの場合）
selected_forms = []
if current_mode == "select":
    st.write("対象とする様式にチェックを入れてください：")
    cols = st.columns(4)
    for i, (form_code, form_label) in enumerate(known_forms.items()):
        with cols[i % 4]:
            if st.checkbox(form_label, key=form_code):
                selected_forms.append(form_code)

st.divider()

# 実行ボタン
if st.button("🚀 全自動処理を開始する", type="primary"):
    
    if current_mode == "select" and not selected_forms:
        st.error("エラー: 「指定様式のみ1本化する」が選ばれましたが、様式が1つもチェックされていません。")
        st.stop()

    progress_bar = st.progress(0)
    status_text = st.empty()
    log_area = st.empty()
    log_messages = []

    def log(msg):
        log_messages.append(msg)
        log_area.text_area("実行ログ", value="\n".join(log_messages), height=300, disabled=True)

    log("="*50)
    log("ダウンロードを開始します...")
    log(f"対象年度: {start_year_str} ～ {end_year_str}")
    log(f"処理モード: {'すべての様式' if current_mode == 'all' else f'指定された様式 ({len(selected_forms)}個)'}")

    with tempfile.TemporaryDirectory() as temp_dir:
        base_save_dir = os.path.join(temp_dir, "processed_data")
        os.makedirs(base_save_dir)
        temp_raw_dir = os.path.join(base_save_dir, "_raw_data_temp")
        os.makedirs(temp_raw_dir)

        try:
            # ==========================================
            # フェーズ1: ダウンロード＆一時解凍
            # ==========================================
            log("\n【フェーズ1】データのダウンロードを開始します...")
            top_url = "https://portal.niad.ac.jp/ptrt/table.html"
            response = requests.get(top_url)
            response.encoding = response.apparent_encoding
            soup = BeautifulSoup(response.text, 'html.parser')
            
            year_links = [a for a in soup.find_all('a', href=True) 
                          if a['href'].endswith('.html') and any(char.isdigit() for char in a['href'])]

            seen_urls = set()
            unique_links = []
            for link in year_links:
                full_url = urljoin(top_url, link.get('href'))
                if full_url not in seen_urls and "table.html" not in full_url:
                    seen_urls.add(full_url)
                    unique_links.append(link)

            # --- 年度によるダウンロード対象の絞り込み ---
            filtered_links = []
            start_y = int(start_year_str) if start_year_str != "すべて" else 0
            end_y = int(end_year_str) if end_year_str != "すべて" else 9999
            
            if start_y > end_y:
                start_y, end_y = end_y, start_y

            for link in unique_links:
                year_text = link.get_text(strip=True).replace('\n', '').replace(' ', '')
                m = re.search(r'^(\d{4})', year_text)
                if m:
                    link_year = int(m.group(1))
                    if start_y <= link_year <= end_y:
                        filtered_links.append(link)
                else:
                    filtered_links.append(link)
            
            unique_links = filtered_links
            # ----------------------------------------------

            total_links = len(unique_links)
            if total_links == 0:
                log("指定された条件に合致するダウンロード可能な年度データが見つかりませんでした。")
                st.stop()

            log(f"{total_links}年度分のデータを取得します。")

            for i, link in enumerate(unique_links):
                progress_val = (i / total_links) * 0.5
                progress_bar.progress(progress_val)
                status_text.text(f"フェーズ1進行中... ({int(progress_val * 100)}%)")

                page_url = urljoin(top_url, link.get('href'))
                year_text = link.get_text(strip=True).replace('\n', '').replace(' ', '')
                safe_year_text = year_text.replace('/', '_').replace(':', '_').replace('（', '(').replace('）', ')')
                year_folder_path = os.path.join(temp_raw_dir, safe_year_text)

                try:
                    page_res = requests.get(page_url)
                    page_res.encoding = page_res.apparent_encoding
                    page_soup = BeautifulSoup(page_res.text, 'html.parser')
                    
                    download_tag = None
                    for a_tag in page_soup.find_all('a', href=True):
                        if "DOWNLOAD" in a_tag.get_text().upper() or a_tag['href'].endswith('.zip'):
                            download_tag = a_tag
                            break
                    
                    if download_tag:
                        zip_url = urljoin(page_url, download_tag.get('href'))
                        log(f" -> ダウンロード＆解凍: {safe_year_text}")
                        zip_res = requests.get(zip_url)
                        with zipfile.ZipFile(io.BytesIO(zip_res.content)) as z:
                            z.extractall(year_folder_path)
                        time.sleep(1)
                    else:
                        log(f"  × スキップ: {safe_year_text}内にデータが見つかりません")
                except Exception as e:
                    log(f"  ! エラー ({safe_year_text}): {e}")

            # ==========================================
            # フェーズ2: 様式別のフォルダ分け
            # ==========================================
            progress_bar.progress(0.5)
            status_text.text("フェーズ2進行中... (50%)")
            log("\n【フェーズ2】ダウンロードしたファイルを様式別に整理します...")
            
            copied_count = 0
            for root, dirs, files in os.walk(temp_raw_dir):
                for file in files:
                    if file.endswith(('.xls', '.xlsx')):
                        file_path = os.path.join(root, file)
                        parts = file.split('_', 1)
                        if len(parts) < 2:
                            continue
                        
                        form_type, _ = os.path.splitext(parts[1])
                        form_type = form_type.replace('-', '_')
                        
                        if current_mode == "select" and form_type not in selected_forms:
                            continue
                        
                        form_dir = os.path.join(base_save_dir, form_type)
                        if not os.path.exists(form_dir):
                            os.makedirs(form_dir)
                            
                        dest_file_path = os.path.join(form_dir, file)
                        shutil.copy2(file_path, dest_file_path)
                        copied_count += 1
            
            log(f" -> {copied_count} 個の対象ファイルを各様式フォルダへ移動しました。")
            shutil.rmtree(temp_raw_dir, ignore_errors=True)

            # ==========================================
            # フェーズ3: データ結合と一本化 (Excel出力対応)
            # ==========================================
            progress_bar.progress(0.6)
            status_text.text("フェーズ3進行中... (60%)")
            log("\n【フェーズ3】各様式フォルダ内でのデータ結合を開始します...")

            form_folders = [d for d in os.listdir(base_save_dir) if os.path.isdir(os.path.join(base_save_dir, d)) and d != "_raw_data_temp"]
            if current_mode == "select":
                form_folders = [d for d in form_folders if d in selected_forms]

            total_forms = len(form_folders)

            if total_forms == 0:
                log("結合対象のフォルダがありませんでした。")
            else:
                for i, form_folder in enumerate(form_folders):
                    progress_val = 0.6 + ((i / total_forms) * 0.4)
                    progress_bar.progress(progress_val)
                    status_text.text(f"フェーズ3進行中... ({int(progress_val * 100)}%)")

                    form_path = os.path.join(base_save_dir, form_folder)
                    log(f"\n■ 処理中: 様式 [{form_folder}]")
                    
                    merged_data = []
                    default_header_idx = 2 
                    
                    # ★修正1: os.listdir の結果を sorted() で並び替える（年度順に処理されるようにする）
                    for file in sorted(os.listdir(form_path)):
                        if file.endswith(('.xls', '.xlsx')):
                            file_path = os.path.join(form_path, file)
                            try:
                                current_header_idx = get_hidden_header_index(file_path)
                                
                                if current_header_idx is not None:
                                    default_header_idx = current_header_idx
                                else:
                                    current_header_idx = default_header_idx

                                df = pd.read_excel(file_path, header=current_header_idx)
                                df.dropna(how='all', axis=0, inplace=True)
                                df.dropna(how='all', axis=1, inplace=True)

                                year = file.split('_')[0]
                                df.insert(0, 'ファイル年度', year)
                                
                                if rename_dict:
                                    df.rename(columns=rename_dict, inplace=True)
                                
                                record_count = len(df)
                                merged_data.append(df)
                                
                                name_part, ext_part = os.path.splitext(file)
                                new_filename = f"{name_part}_({record_count}_records){ext_part}"
                                os.rename(file_path, os.path.join(form_path, new_filename))
                                log(f"    - 読込成功: {new_filename} (非表示ヘッダー: {current_header_idx + 1}行目)")
                            except Exception as e:
                                log(f"    × 読込エラー: {file} - {e}")
                                
                    if merged_data:
                        final_df = pd.concat(merged_data, ignore_index=True)

                        # ★修正2: すべて合体させた後に、念のため「ファイル年度」で並べ替え（ソート）を確実に行う
                        # kind='stable' を指定することで、同一年度内での元の行の並び順（大学の並び順など）を崩さずに年度順にソートします
                        final_df = final_df.sort_values(by='ファイル年度', ascending=True, kind='stable', ignore_index=True)

                        # --- 学校名と大学名の統合 ---
                        has_gakko = '学校名' in final_df.columns
                        has_daigaku = '大学名' in final_df.columns
                        
                        if has_gakko and has_daigaku:
                            combined_col = final_df['学校名'].fillna(final_df['大学名'])
                            idx_gakko = list(final_df.columns).index('学校名')
                            idx_daigaku = list(final_df.columns).index('大学名')
                            insert_idx = min(idx_gakko, idx_daigaku)
                            
                            final_df = final_df.drop(columns=['学校名', '大学名'])
                            final_df.insert(insert_idx, '学校名・大学名', combined_col)
                            log(f"    - 「学校名」と「大学名」列を「学校名・大学名」に統合しました。")
                            
                        elif has_gakko:
                            final_df = final_df.rename(columns={'学校名': '学校名・大学名'})
                            log(f"    - 「学校名」列を「学校名・大学名」に変更しました。")
                        elif has_daigaku:
                            final_df = final_df.rename(columns={'大学名': '学校名・大学名'})
                            log(f"    - 「大学名」列を「学校名・大学名」に変更しました。")

                        # --- 途中追加・途中終了列の検知とアラート（ピンク色） ---
                        pink_columns = []
                        if not final_df.empty and 'ファイル年度' in final_df.columns:
                            file_years = final_df['ファイル年度'].dropna().astype(str)
                            min_year = file_years.min()
                            max_year = file_years.max()
                            tmp_rename_mapping = {}
                            
                            for col in final_df.columns:
                                if col == 'ファイル年度':
                                    continue
                                
                                valid_mask = final_df[col].notna() & (final_df[col].astype(str).str.strip() != '') & (final_df[col].astype(str).str.lower() != 'nan')
                                valid_data = final_df[valid_mask]
                                
                                if not valid_data.empty:
                                    first_year = valid_data['ファイル年度'].astype(str).min()
                                    last_year = valid_data['ファイル年度'].astype(str).max()
                                    
                                    suffix = ""
                                    if first_year > min_year and last_year < max_year:
                                        suffix = f"（{first_year}年度～{last_year}年度）"
                                    elif first_year > min_year:
                                        suffix = f"（{first_year}年度～）"
                                    elif last_year < max_year:
                                        suffix = f"（～{last_year}年度）"
                                    
                                    if suffix:
                                        new_col_name = f"{col}{suffix}"
                                        tmp_rename_mapping[col] = new_col_name
                                        pink_columns.append(new_col_name) 
                                        
                            if tmp_rename_mapping:
                                final_df.rename(columns=tmp_rename_mapping, inplace=True)
                                log(f"    ! 注意: {len(tmp_rename_mapping)}個の列が「途中追加」または「途中終了」していることを検知しました。")

                        final_record_count = len(final_df)
                        # ★拡張子を .csv から .xlsx に変更して保存
                        merged_filename = f"{form_folder}_merged_({final_record_count}_records).xlsx"
                        save_path = os.path.join(form_path, merged_filename)
                        
                        final_df.to_excel(save_path, index=False, engine='openpyxl')

                        # ★ピンク色を塗る対象があれば、開いて色を塗る
                        if pink_columns:
                            try:
                                wb_merged = openpyxl.load_workbook(save_path)
                                ws_merged = wb_merged.active
                                pink_fill = PatternFill(patternType='solid', fgColor='FFC0CB')
                                
                                for cell in ws_merged[1]:
                                    if cell.value in pink_columns:
                                        cell.fill = pink_fill
                                
                                wb_merged.save(save_path)
                            except Exception as e:
                                log(f"    ! 色の適用中にエラーが発生しました: {e}")

                        log(f"  ★ 結合完了: {merged_filename} を生成しました")

            with open(os.path.join(base_save_dir, "run_log.txt"), "w", encoding="utf-8") as f:
                f.write("\n".join(log_messages))

            # ==========================================
            # フェーズ4: ZIP化とダウンロードボタンの生成
            # ==========================================
            progress_bar.progress(1.0)
            status_text.text("処理完了 (100%)")
            log("\n==================================================")
            log("すべての処理が完了しました！ZIPファイルを作成しています...")

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for root, dirs, files in os.walk(base_save_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, base_save_dir)
                        zip_file.write(file_path, arcname)
            
            st.success("✅ 全ての処理が完了しました！下のボタンから結果をダウンロードしてください。")
            
            st.download_button(
                label="📁 整理済みのデータをダウンロード (ZIP)",
                data=zip_buffer.getvalue(),
                file_name=f"school_basic_survey_data_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                mime="application/zip",
                type="primary"
            )

        except Exception as e:
            st.error(f"\n予期せぬエラーが発生しました: {e}")
            log(f"エラー詳細: {e}")
